#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import openpyxl
from openpyxl.cell import MergedCell
from datetime import datetime


class JiraExcel(object):

    COL_PROJECT = 1
    COL_FEATURE_ID = 2
    COL_FEATURE_NAME = 3
    COL_STORY_ID = 4
    COL_STORY_TITLE = 5
    COL_STORY_POINT = 6
    COL_STORY_DESCRIPTION = 8
    COL_STORY_ACCEPTANCE_CRITERIA = 9
    COL_STORY_ACTION = 10
    COL_STORY_STATUS = 11
    COL_STORY_COMPONENTS = 12
    COL_STORY_PRODUCTS = 13
    COL_STORY_FIX_VERSIONS = 14

    def __init__(self, excel_file: str):
        self.__excel_file = excel_file
        self.__workbook = None
        self.__sheet = None

    def open(self):
        self.__workbook = openpyxl.load_workbook(self.__excel_file)
        self.__sheet = self.__workbook.worksheets[0]

    def get_items(self) -> list:
        result = []
        for row in range(1, self.__sheet.max_row + 1):
            if row == 1:
                continue
            item = JiraExcelItem()
            item.row = row
            item.project = self.get_cell(row, JiraExcel.COL_PROJECT).value
            item.feature_id = self.get_cell(row, JiraExcel.COL_FEATURE_ID).value
            item.feature_name = self.get_cell(row, JiraExcel.COL_FEATURE_NAME).value
            item.story_id = self.get_cell(row, JiraExcel.COL_STORY_ID).value
            item.story_title = self.get_cell(row, JiraExcel.COL_STORY_TITLE).value
            item.story_point = self.get_cell(row, JiraExcel.COL_STORY_POINT).value
            item.story_description = self.get_cell(row, JiraExcel.COL_STORY_DESCRIPTION).value
            item.story_acceptance_criteria = self.get_cell(row, JiraExcel.COL_STORY_ACCEPTANCE_CRITERIA).value
            item.story_action = self.get_cell(row, JiraExcel.COL_STORY_ACTION).value
            item.story_status = self.get_cell(row, JiraExcel.COL_STORY_STATUS).value
            item.story_components = self.get_cell(row, JiraExcel.COL_STORY_COMPONENTS).value
            item.story_products = self.get_cell(row, JiraExcel.COL_STORY_PRODUCTS).value
            item.story_fix_versions = self.get_cell(row, JiraExcel.COL_STORY_FIX_VERSIONS).value
            result.append(item)
        return result

    def set_item(self, item):
        self.get_cell(item.row, JiraExcel.COL_PROJECT).value = item.project
        self.get_cell(item.row, JiraExcel.COL_FEATURE_ID).value = item.feature_id
        self.get_cell(item.row, JiraExcel.COL_FEATURE_NAME).value = item.feature_name
        self.get_cell(item.row, JiraExcel.COL_STORY_ID).value = item.story_id
        self.get_cell(item.row, JiraExcel.COL_STORY_TITLE).value = item.story_title
        self.get_cell(item.row, JiraExcel.COL_STORY_POINT).value = item.story_point
        self.get_cell(item.row, JiraExcel.COL_STORY_DESCRIPTION).value = item.story_description
        self.get_cell(item.row, JiraExcel.COL_STORY_ACCEPTANCE_CRITERIA).value = item.story_acceptance_criteria
        # self.get_cell(item.row, JiraExcel.COL_STORY_ACTION).value = item.story_action
        self.get_cell(item.row, JiraExcel.COL_STORY_STATUS).value = item.story_status
        self.get_cell(item.row, JiraExcel.COL_STORY_COMPONENTS).value = item.story_components
        self.get_cell(item.row, JiraExcel.COL_STORY_PRODUCTS).value = item.story_products
        self.get_cell(item.row, JiraExcel.COL_STORY_FIX_VERSIONS).value = item.story_fix_versions

    def get_cell(self, row: int, col: int):
        cell = self.__sheet.cell(row, col)
        if isinstance(cell, MergedCell):
            for merged_cell_range in self.__sheet.merged_cell_ranges:
                if cell.coordinate in merged_cell_range:
                    cell = self.__sheet.cell(merged_cell_range.min_row, merged_cell_range.min_col)
                    break
        return cell

    def save(self, new_file_name=None):
        if new_file_name is None or new_file_name == "":
            self.__workbook.save(self.__excel_file)
        else:
            self.__workbook.save(new_file_name)

    def backup(self):
        date = datetime.now()
        date_str = date.strftime("%Y%m%d_%H%M%S")
        file_paths = os.path.splitext(self.__excel_file)
        backup_file = file_paths[0] + "_Backup_" + date_str + file_paths[1]
        self.save(backup_file)

    def close(self):
        self.__workbook.close()
        self.__workbook = None
        self.__sheet = None

    def show_items_details(self):
        items = self.get_items()
        for i in range(len(items)):
            print(items[i])


class JiraExcelItem(object):

    def __init__(self):
        self.__row = None
        self.__project = None
        self.__feature_id = None
        self.__feature_name = None
        self.__story_id = None
        self.__story_title = None
        self.__story_point = None
        self.__story_description = None
        self.__story_acceptance_criteria = None
        self.__story_action = None
        self.__story_status = None
        self.__story_components = None
        self.__story_products = None
        self.__story_fix_versions = None

    def __str__(self):
        return '''
            row: %s,
            project: %s,
            feature_id: %s,
            feature_name: %s,
            story_id: %s,
            story_title: %s,
            story_point: %s,
            story_description: %s,
            story_acceptance_criteria: %s,
            story_action: %s,
            story_status: %s,
            story_components: %s,
            story_products: %s,
            story_fix_versions: %s
        ''' % (
            self.__row,
            self.__project,
            self.__feature_id,
            self.__feature_name,
            self.__story_id,
            self.__story_title,
            self.__story_point,
            self.__story_description,
            self.__story_acceptance_criteria,
            self.__story_action,
            self.__story_status,
            self.__story_components,
            self.__story_products,
            self.__story_fix_versions
        )

    @property
    def row(self):
        return self.__row

    @row.setter
    def row(self, value):
        self.__row = value

    @property
    def project(self):
        return self.__project

    @project.setter
    def project(self, value):
        self.__project = JiraExcelItem.trim_str_value(value)

    @property
    def feature_id(self):
        return self.__feature_id

    @feature_id.setter
    def feature_id(self, value):
        self.__feature_id = JiraExcelItem.trim_str_value(value)

    @property
    def feature_name(self):
        return self.__feature_name

    @feature_name.setter
    def feature_name(self, value):
        self.__feature_name = JiraExcelItem.trim_str_value(value)

    @property
    def story_id(self):
        return self.__story_id

    @story_id.setter
    def story_id(self, value):
        self.__story_id = JiraExcelItem.trim_str_value(value)

    @property
    def story_title(self):
        return self.__story_title

    @story_title.setter
    def story_title(self, value):
        self.__story_title = JiraExcelItem.trim_str_value(value)

    @property
    def story_point(self):
        return self.__story_point

    @story_point.setter
    def story_point(self, value):
        self.__story_point = value

    @property
    def story_description(self):
        return self.__story_description

    @story_description.setter
    def story_description(self, value):
        self.__story_description = JiraExcelItem.trim_str_value(value)

    @property
    def story_acceptance_criteria(self):
        return self.__story_acceptance_criteria

    @story_acceptance_criteria.setter
    def story_acceptance_criteria(self, value):
        self.__story_acceptance_criteria = JiraExcelItem.trim_str_value(value)

    @property
    def story_action(self):
        return self.__story_action

    @story_action.setter
    def story_action(self, value):
        self.__story_action = JiraExcelItem.trim_str_value(value)

    @property
    def story_status(self):
        return self.__story_status

    @story_status.setter
    def story_status(self, value):
        self.__story_status = JiraExcelItem.trim_str_value(value)

    @property
    def story_components(self):
        return self.__story_components

    @story_components.setter
    def story_components(self, value):
        self.__story_components = JiraExcelItem.trim_str_value(value)

    @property
    def story_products(self):
        return self.__story_products

    @story_products.setter
    def story_products(self, value):
        self.__story_products = JiraExcelItem.trim_str_value(value)

    @property
    def story_fix_versions(self):
        return self.__story_fix_versions

    @story_fix_versions.setter
    def story_fix_versions(self, value):
        self.__story_fix_versions = JiraExcelItem.trim_str_value(value)

    @staticmethod
    def trim_str_value(value):
        return None if value is None or value.strip() == "" else value.strip()
